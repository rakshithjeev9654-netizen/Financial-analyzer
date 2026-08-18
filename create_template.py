"""
create_template.py
Generates the downloadable Excel template for the Financial Report Analyzer.
Run once to create: static/financial_template.xlsx
"""

import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from pathlib import Path

OUTPUT = Path("static/financial_template.xlsx")
OUTPUT.parent.mkdir(parents=True, exist_ok=True)


def thin_border():
    s = Side(style="thin", color="D1D5DB")
    return Border(left=s, right=s, top=s, bottom=s)


def make_template():
    wb = openpyxl.Workbook()

    # ── Sheet 1: Data Entry ──────────────────────────────────────────
    ws = wb.active
    ws.title = "Financial Data"

    # Color palette
    NAVY_FILL   = PatternFill("solid", fgColor="0F2554")
    BLUE_FILL   = PatternFill("solid", fgColor="1A56DB")
    GREEN_FILL  = PatternFill("solid", fgColor="059669")
    PURPLE_FILL = PatternFill("solid", fgColor="7C3AED")
    TEAL_FILL   = PatternFill("solid", fgColor="0891B2")
    GRAY_FILL   = PatternFill("solid", fgColor="F8FAFC")
    LIGHT_BLUE  = PatternFill("solid", fgColor="EFF6FF")
    LIGHT_GREEN = PatternFill("solid", fgColor="F0FDF4")
    LIGHT_PURP  = PatternFill("solid", fgColor="FAF5FF")
    LIGHT_TEAL  = PatternFill("solid", fgColor="F0FDFA")
    YELLOW_FILL = PatternFill("solid", fgColor="FEFCE8")

    WHITE_FONT  = Font(color="FFFFFF", bold=True, name="Segoe UI", size=11)
    BOLD_NAVY   = Font(color="0F2554", bold=True, name="Segoe UI", size=10)
    NORM_FONT   = Font(name="Segoe UI", size=10)
    SMALL_FONT  = Font(name="Segoe UI", size=9, italic=True, color="6B7280")
    CENTER      = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT        = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    RIGHT_ALIGN = Alignment(horizontal="right",  vertical="center")

    # ── Title row ────────────────────────────────────────────────────
    ws.merge_cells("A1:P1")
    ws["A1"] = "Financial Report Analyzer — Data Entry Template"
    ws["A1"].fill      = NAVY_FILL
    ws["A1"].font      = Font(color="FFFFFF", bold=True, name="Segoe UI", size=14)
    ws["A1"].alignment = CENTER
    ws.row_dimensions[1].height = 36

    # ── Subtitle ─────────────────────────────────────────────────────
    ws.merge_cells("A2:P2")
    ws["A2"] = "Fill in the yellow cells. Each row = one reporting period. Delete example rows before uploading."
    ws["A2"].fill      = YELLOW_FILL
    ws["A2"].font      = Font(color="92400E", bold=True, name="Segoe UI", size=10)
    ws["A2"].alignment = CENTER
    ws.row_dimensions[2].height = 22

    # ── Section headers row 3 ─────────────────────────────────────────
    sections = [
        ("A3", "B3", "PERIOD",           NAVY_FILL),
        ("B3", "F3", "INCOME STATEMENT", BLUE_FILL),
        ("F3", "J3", "BALANCE SHEET — ASSETS", GREEN_FILL),
        ("J3", "M3", "BALANCE SHEET — LIABILITIES & EQUITY", PURPLE_FILL),
        ("M3", "P3", "WORKING CAPITAL",  TEAL_FILL),
    ]
    # Merge and style section headers
    section_ranges = [
        ("A3:A3",  "PERIOD",                         NAVY_FILL),
        ("B3:E3",  "INCOME STATEMENT",               BLUE_FILL),
        ("F3:I3",  "BALANCE SHEET — ASSETS",         GREEN_FILL),
        ("J3:L3",  "BALANCE SHEET — LIAB. & EQUITY", PURPLE_FILL),
        ("M3:P3",  "WORKING CAPITAL",                TEAL_FILL),
    ]
    for rng, label, fill in section_ranges:
        ws.merge_cells(rng)
        cell = ws[rng.split(":")[0]]
        cell.value     = label
        cell.fill      = fill
        cell.font      = WHITE_FONT
        cell.alignment = CENTER
    ws.row_dimensions[3].height = 22

    # ── Column headers row 4 ─────────────────────────────────────────
    headers = [
        # col,  label,                    fill,         note
        ("A", "period",                   GRAY_FILL,    "e.g. FY2024, 2024, Q1 2024"),
        ("B", "revenue",                  LIGHT_BLUE,   "Total revenue / sales / turnover"),
        ("C", "cogs",                     LIGHT_BLUE,   "Cost of goods sold / cost of sales"),
        ("D", "operating_income",         LIGHT_BLUE,   "EBIT / operating profit"),
        ("E", "net_income",               LIGHT_BLUE,   "Net profit / profit after tax"),
        ("F", "current_assets",           LIGHT_GREEN,  "Total current assets"),
        ("G", "cash",                     LIGHT_GREEN,  "Cash and cash equivalents"),
        ("H", "inventory",                LIGHT_GREEN,  "Inventory / stock"),
        ("I", "total_assets",             LIGHT_GREEN,  "Total assets (balance sheet total)"),
        ("J", "current_liabilities",      LIGHT_PURP,   "Total current liabilities"),
        ("K", "total_liabilities",        LIGHT_PURP,   "Total liabilities"),
        ("L", "shareholders_equity",      LIGHT_PURP,   "Total equity / net assets"),
        ("M", "accounts_receivable",      LIGHT_TEAL,   "Trade receivables / debtors"),
        ("N", "accounts_payable",         LIGHT_TEAL,   "Trade payables / creditors"),
        ("O", "interest_expense",         LIGHT_TEAL,   "Interest paid / finance costs"),
        ("P", "extra_notes",              GRAY_FILL,    "Optional: any notes (ignored by analyzer)"),
    ]

    for col_letter, label, fill, note in headers:
        cell = ws[f"{col_letter}4"]
        cell.value     = label
        cell.fill      = fill
        cell.font      = BOLD_NAVY
        cell.alignment = CENTER
        cell.border    = thin_border()

    ws.row_dimensions[4].height = 30

    # ── Tooltip / hint row 5 ─────────────────────────────────────────
    for col_letter, label, fill, note in headers:
        cell = ws[f"{col_letter}5"]
        cell.value     = note
        cell.fill      = fill
        cell.font      = SMALL_FONT
        cell.alignment = CENTER
        cell.border    = thin_border()
    ws.row_dimensions[5].height = 28

    # ── Example data rows 6–8 ────────────────────────────────────────
    example_data = [
        ["FY2022", 4200000, 2520000, 588000,  420000,  980000,  210000, 310000, 2100000, 490000,  1050000, 1050000, 380000, 210000, 84000,  "Example row — delete before uploading"],
        ["FY2023", 4830000, 2800000, 724500,  530000,  1150000, 260000, 345000, 2420000, 540000,  1100000, 1320000, 430000, 225000, 88000,  ""],
        ["FY2024", 5460000, 3100000, 873600,  654000,  1320000, 310000, 360000, 2740000, 580000,  1150000, 1590000, 490000, 240000, 91000,  ""],
    ]

    EXAMPLE_FILL = PatternFill("solid", fgColor="FFFBEB")
    NUM_FONT     = Font(name="Segoe UI", size=10, color="374151")

    for row_idx, row_data in enumerate(example_data, start=6):
        ws.row_dimensions[row_idx].height = 22
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill   = EXAMPLE_FILL
            cell.font   = NUM_FONT
            cell.border = thin_border()
            if col_idx == 1:
                cell.alignment = CENTER
                cell.font = Font(name="Segoe UI", size=10, bold=True, color="374151")
            elif isinstance(value, (int, float)):
                cell.alignment  = RIGHT_ALIGN
                cell.number_format = '#,##0'
            else:
                cell.alignment = LEFT

    # ── Empty data rows 9–18 for user input ──────────────────────────
    USER_FILL = PatternFill("solid", fgColor="FFFFFF")
    for row_idx in range(9, 19):
        ws.row_dimensions[row_idx].height = 22
        for col_idx in range(1, 17):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill   = USER_FILL if col_idx > 1 else GRAY_FILL
            cell.border = thin_border()
            cell.font   = NORM_FONT
            if col_idx > 1 and col_idx < 16:
                cell.number_format = '#,##0'
                cell.alignment     = RIGHT_ALIGN
            else:
                cell.alignment = CENTER

    # ── Column widths ────────────────────────────────────────────────
    col_widths = {
        "A": 14, "B": 16, "C": 16, "D": 18, "E": 14,
        "F": 16, "G": 12, "H": 12, "I": 16,
        "J": 20, "K": 18, "L": 20,
        "M": 20, "N": 18, "O": 18, "P": 38,
    }
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    # ── Freeze panes ─────────────────────────────────────────────────
    ws.freeze_panes = "B6"

    # ── Sheet 2: Instructions ─────────────────────────────────────────
    wi = wb.create_sheet("Instructions")

    def inst_row(row, text, style="normal"):
        wi.merge_cells(f"A{row}:F{row}")
        c = wi[f"A{row}"]
        c.value = text
        if style == "title":
            c.fill = NAVY_FILL
            c.font = Font(color="FFFFFF", bold=True, name="Segoe UI", size=14)
            c.alignment = CENTER
            wi.row_dimensions[row].height = 36
        elif style == "heading":
            c.fill = BLUE_FILL
            c.font = Font(color="FFFFFF", bold=True, name="Segoe UI", size=11)
            c.alignment = LEFT
            wi.row_dimensions[row].height = 24
        elif style == "subhead":
            c.fill = LIGHT_BLUE
            c.font = Font(color="0F2554", bold=True, name="Segoe UI", size=10)
            c.alignment = LEFT
            wi.row_dimensions[row].height = 20
        elif style == "note":
            c.fill = YELLOW_FILL
            c.font = Font(color="92400E", name="Segoe UI", size=10)
            c.alignment = LEFT
            wi.row_dimensions[row].height = 18
        else:
            c.font = Font(name="Segoe UI", size=10, color="374151")
            c.alignment = LEFT
            wi.row_dimensions[row].height = 18

    inst_row(1,  "Financial Report Analyzer — How to Use This Template", "title")
    inst_row(2,  "")
    inst_row(3,  "STEP 1 — Fill in your financial data", "heading")
    inst_row(4,  "  Go to the 'Financial Data' sheet.", "normal")
    inst_row(5,  "  Each ROW = one reporting period (e.g. FY2022, FY2023, FY2024).", "normal")
    inst_row(6,  "  Each COLUMN = one financial line item (revenue, costs, assets etc.).", "normal")
    inst_row(7,  "  Delete the 3 example rows (rows 6-8) before uploading.", "note")
    inst_row(8,  "")
    inst_row(9,  "STEP 2 — Enter numbers as plain numbers", "heading")
    inst_row(10, "  Enter numbers without currency symbols: 5000000 not $5,000,000", "normal")
    inst_row(11, "  Comma separators are OK: 5,000,000", "normal")
    inst_row(12, "  Use negative numbers for losses: -250000", "normal")
    inst_row(13, "  All numbers should be in the SAME unit (e.g. all in USD, all full values).", "note")
    inst_row(14, "")
    inst_row(15, "STEP 3 — Save and upload", "heading")
    inst_row(16, "  Save as .xlsx or .csv format.", "normal")
    inst_row(17, "  Go to the Financial Report Analyzer upload page.", "normal")
    inst_row(18, "  Drag and drop or browse to your file.", "normal")
    inst_row(19, "  Click 'Analyze Report'.", "normal")
    inst_row(20, "")
    inst_row(21, "COLUMN REFERENCE", "heading")
    inst_row(22, "  Column names in the template are the EXACT names the analyzer looks for.", "normal")
    inst_row(23, "  The analyzer also recognizes common alternatives (see below).", "normal")
    inst_row(24, "")

    # Column reference table
    ref_data = [
        ("Template Column",     "Also recognized as",                                     "Required for"),
        ("period",              "Year, Date, Fiscal Year, Quarter",                        "Period labels on charts"),
        ("revenue",             "Sales, Turnover, Net Revenue, Total Revenue",             "All profitability ratios"),
        ("cogs",                "Cost of Sales, Cost of Revenue, Direct Costs",            "Gross profit margin"),
        ("operating_income",    "EBIT, Operating Profit, Op Income",                      "Operating margin, Interest coverage"),
        ("net_income",          "Net Profit, PAT, Profit After Tax, Earnings",             "Net margin, ROA, ROE"),
        ("current_assets",      "Total Current Assets, Short-Term Assets, CA",            "Current ratio, Quick ratio"),
        ("cash",                "Cash & Cash Equivalents, Cash Balance",                  "Cash ratio"),
        ("inventory",           "Inventories, Stock, Merchandise",                        "Quick ratio, Inventory turnover"),
        ("total_assets",        "Assets Total, TA, Balance Sheet Total",                  "ROA, Asset turnover, Debt ratio"),
        ("current_liabilities", "Total Current Liabilities, Short-Term Liabilities, CL",  "Current ratio, Quick ratio"),
        ("total_liabilities",   "Total Debt, Liabilities Total, TL",                     "Debt ratio, D/E ratio"),
        ("shareholders_equity", "Total Equity, Net Assets, Book Value, SE",              "D/E ratio, ROE, Equity ratio"),
        ("accounts_receivable", "Receivables, Trade Receivables, Debtors, AR",           "Receivables turnover, Days"),
        ("accounts_payable",    "Payables, Trade Payables, Creditors, AP",               "Payables turnover, CCC"),
        ("interest_expense",    "Finance Costs, Interest Paid, Borrowing Costs, IE",     "Interest coverage ratio"),
    ]

    HDR_FILL = PatternFill("solid", fgColor="0F2554")
    ROW_FILL = [PatternFill("solid", fgColor="F8FAFC"), PatternFill("solid", fgColor="FFFFFF")]

    for ridx, (col1, col2, col3) in enumerate(ref_data, start=25):
        wi.row_dimensions[ridx].height = 20
        for cidx, val in enumerate([col1, col2, col3], start=1):
            c = wi.cell(row=ridx, column=cidx, value=val)
            c.border = thin_border()
            c.alignment = LEFT
            if ridx == 25:
                c.fill = HDR_FILL
                c.font = Font(color="FFFFFF", bold=True, name="Segoe UI", size=10)
            else:
                c.fill = ROW_FILL[(ridx - 26) % 2]
                c.font = Font(name="Segoe UI", size=9,
                              bold=(cidx == 1), color="374151")

    wi.column_dimensions["A"].width = 22
    wi.column_dimensions["B"].width = 50
    wi.column_dimensions["C"].width = 38
    for col in ["D", "E", "F"]:
        wi.column_dimensions[col].width = 12

    wb.save(OUTPUT)
    print(f"Template saved: {OUTPUT}")


if __name__ == "__main__":
    make_template()
